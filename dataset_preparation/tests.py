from io import BytesIO

import requests
from django.contrib.auth.models import User
from django.test import SimpleTestCase, TestCase
from mock import Mock, patch
from openpyxl import Workbook

from budgetportal import models as budgetportal_models
from budgetportal.dataset_uploading.dataset_preprocessor import (
    AENE_HEADERS,
    BUDGET_ACTUAL_HEADERS,
    ENE_HEADERS,
    EPRE_HEADERS,
)

from .models import DatasetImportJob, DatasetPreparationJob
from .services import (
    download_source_file,
    find_header_row,
    iter_sheet_records,
    load_rows_from_excel,
    build_consolidated_rows,
    transform_consolidated_records,
    transform_records,
)
from .tasks import run_dataset_import_job, run_dataset_preparation_job


def build_source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Report title"])
    worksheet.append(
        [
            "Province",
            "Vote No.",
            "Department",
            "Programme No.",
            "Programme",
            "Subprogramme No.",
            "Subprogramme",
            "Econ1",
            "Econ2",
            "Econ3",
            "Econ4",
            "Econ5",
            "Function group",
            "Function group 2",
            "2026/27 Revised baseline",
            "2026/27 Adjusted appropriation",
        ]
    )
    worksheet.append(
        [
            "KwaZulu-Natal",
            "Vote 6",
            "education",
            1,
            "Administration",
            1,
            "Office of the MEC",
            "Current payments",
            "Compensation of employees",
            "Salaries and wages",
            "Level 4",
            "Level 5",
            "Learning and culture",
            "Basic education",
            10,
            20,
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_national_source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Report title"])
    worksheet.append(
        [
            "Vote No.",
            "Department",
            "Programme No.",
            "Programme",
            "Subprogramme No.",
            "Subprogramme",
            "Econ1",
            "Econ2",
            "Econ3",
            "Econ4",
            "Econ5",
            "Function group",
            "2022 Audited outcome",
            "2023 Audited outcome",
            "2024 Audited outcome",
            "2025 Adjusted appropriation",
            "2025 Revised estimate",
            "2025 Voted (Main appropriation)",
            "2026 Budget",
            "2027 MTEF1",
            "2028 MTEF2",
        ]
    )
    worksheet.append(
        [
            "Vote 1",
            "the presidency",
            1,
            "Administration",
            1,
            "Ministry",
            "Current payments",
            "Compensation of employees",
            "Salaries and wages",
            "Level 4",
            "Level 5",
            "Executive and administration",
            10,
            20,
            30,
            40,
            50,
            60,
            70,
            80,
            90,
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_consolidated_source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(
        [
            "BudgetYear",
            "Type",
            "Function group",
            "Budget group",
            "Econ2",
            "Econ3",
            "FinYear",
            "Value",
        ]
    )
    worksheet.append(
        [
            "Budget 2026",
            "Expenditure",
            "Executive and administration",
            "Executive and administration",
            "Compensation of employees",
            "Salaries and wages",
            "2026/27",
            100,
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_aene_tabular_source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data (Tabular)"
    worksheet.append(
        [
            "Voted_DirectCharges",
            "Vote No.",
            "Department",
            "Programme No.",
            "Programme",
            "Subprogramme No.",
            "Subprogramme",
            "Econ1",
            "Econ2",
            "Econ3",
            "Econ4",
            "Econ5",
            "FinYear",
            "FY description",
            "Value",
        ]
    )
    worksheet.append(
        [
            "Voted",
            "01",
            "The Presidency",
            "01",
            "Administration",
            "01",
            "Ministry",
            "Current",
            "Transfers and subsidies",
            "Households",
            "Social benefits",
            "Social benefits",
            "2025/26",
            "Adjusted appropriation",
            100,
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_aene_flat_source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data (Flat)"
    worksheet.append(
        [
            "Voted_DirectCharges",
            "Vote No.",
            "Department",
            "Programme No.",
            "Programme",
            "Subprogramme No.",
            "Subprogramme",
            "Econ1",
            "Econ2",
            "Econ3",
            "Econ4",
            "Econ5",
            "2025/26 - Appropriation",
            "2025/26 - Adjusted appropriation",
        ]
    )
    worksheet.append(
        [
            "Voted",
            "01",
            "The Presidency",
            "01",
            "Administration",
            "01",
            "Ministry",
            "Current",
            "Transfers and subsidies",
            "Households",
            "Social benefits",
            "Social benefits",
            0,
            100,
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class DatasetPreparationServiceTest(SimpleTestCase):
    @patch("dataset_preparation.services.time.sleep")
    @patch("dataset_preparation.services.requests.get")
    def test_download_source_file_retries_transient_request_failures(
        self, mock_get, mock_sleep
    ):
        mock_response = Mock()
        mock_response.content = b"file-bytes"
        mock_response.raise_for_status = Mock()
        mock_get.side_effect = [
            requests.exceptions.ConnectionError("temporary dns failure"),
            mock_response,
        ]

        content = download_source_file("https://example.com/epre.xlsx")

        self.assertEqual(content, b"file-bytes")
        self.assertEqual(mock_get.call_count, 2)
        mock_sleep.assert_called_once()

    @patch("dataset_preparation.services.time.sleep")
    @patch("dataset_preparation.services.requests.get")
    def test_download_source_file_raises_after_retry_limit(self, mock_get, mock_sleep):
        mock_get.side_effect = requests.exceptions.ConnectionError(
            "temporary dns failure"
        )

        with self.assertRaises(requests.exceptions.ConnectionError):
            download_source_file("https://example.com/epre.xlsx")

        self.assertEqual(mock_get.call_count, 4)
        self.assertEqual(mock_sleep.call_count, 3)

    def test_transforms_epre_rows_into_expected_schema(self):
        workbook_bytes = build_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data")
        header_row_index = find_header_row(rows)
        records = list(iter_sheet_records(rows, header_row_index))

        financial_year = Mock(slug="2026-27")
        transformed_rows = transform_records(
            records,
            financial_year,
            DatasetPreparationJob.DATASET_TYPE_EPRE,
        )

        self.assertEqual(list(transformed_rows[0].keys()), EPRE_HEADERS)
        self.assertEqual(transformed_rows[0]["Government"], "KwaZulu-Natal")
        self.assertEqual(transformed_rows[0]["VoteNumber"], 6)
        self.assertEqual(transformed_rows[0]["BudgetPhase"], "Main appropriation")
        self.assertEqual(transformed_rows[0]["Value"], 10000)

    def test_transforms_budget_vs_actual_rows_into_expected_schema(self):
        workbook_bytes = build_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data")
        header_row_index = find_header_row(rows)
        records = list(iter_sheet_records(rows, header_row_index))

        financial_year = Mock(slug="2026-27")
        transformed_rows = transform_records(
            records,
            financial_year,
            "Budget-vs-Actual-Provincial",
        )

        adjusted_row = [
            row for row in transformed_rows
            if row["BudgetPhase"] == "Adjusted appropriation"
        ][0]

        self.assertEqual(list(adjusted_row.keys()), BUDGET_ACTUAL_HEADERS)
        self.assertEqual(adjusted_row["AmountKind"], "Total")
        self.assertEqual(adjusted_row["Value"], 20000)

    def test_transforms_ene_rows_into_expected_schema(self):
        workbook_bytes = build_national_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data")
        header_row_index = find_header_row(rows)
        records = list(iter_sheet_records(rows, header_row_index))

        financial_year = Mock(slug="2026-27")
        transformed_rows = transform_records(
            records,
            financial_year,
            DatasetPreparationJob.DATASET_TYPE_ENE,
        )

        self.assertEqual(list(transformed_rows[0].keys()), ENE_HEADERS)
        self.assertEqual(transformed_rows[0]["VoteNumber"], 1)
        self.assertEqual(transformed_rows[0]["Department"], "The Presidency")
        self.assertEqual(transformed_rows[0]["BudgetPhase"], "Main appropriation")
        self.assertEqual(transformed_rows[0]["FinancialYear"], "2026")
        self.assertEqual(transformed_rows[0]["Value"], 70000)

    def test_aggregates_ene_rows_into_consolidated_schema(self):
        ene_rows = [
            {
                "FunctionGroup1": "Executive and administration",
                "EconomicClassification2": "Compensation of employees",
                "EconomicClassification3": "Salaries and wages",
                "FinancialYear": "2026",
                "Value": 70000,
            },
            {
                "FunctionGroup1": "Executive and administration",
                "EconomicClassification2": "Compensation of employees",
                "EconomicClassification3": "Salaries and wages",
                "FinancialYear": "2026",
                "Value": 30000,
            },
        ]

        rows = build_consolidated_rows(ene_rows)

        self.assertEqual(len(rows), 1)
        self.assertEqual(
            list(rows[0].keys()),
            [
                "FunctionGroup",
                "EconomicClassification2",
                "EconomicClassification3",
                "FinancialYear",
                "Value",
            ],
        )
        self.assertEqual(rows[0]["Value"], 100000)

    def test_transforms_consolidated_account_rows_into_expected_schema(self):
        workbook_bytes = build_consolidated_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data")
        records = list(iter_sheet_records(rows, 0))

        transformed_rows = transform_consolidated_records(
            records,
            Mock(slug="2026-27"),
        )

        self.assertEqual(len(transformed_rows), 1)
        self.assertEqual(transformed_rows[0]["FunctionGroup"], "Executive and administration")
        self.assertEqual(transformed_rows[0]["FinancialYear"], "2026")
        self.assertEqual(transformed_rows[0]["Value"], 100000)

    def test_transforms_aene_rows_into_expected_schema(self):
        workbook_bytes = build_national_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data")
        header_row_index = find_header_row(rows)
        records = list(iter_sheet_records(rows, header_row_index))

        financial_year = Mock(slug="2026-27")
        transformed_rows = transform_records(
            records,
            financial_year,
            DatasetPreparationJob.DATASET_TYPE_AENE,
        )

        self.assertEqual(list(transformed_rows[0].keys()), AENE_HEADERS)
        self.assertEqual(transformed_rows[0]["VoteNumber"], 1)
        self.assertEqual(transformed_rows[0]["BudgetPhase"], "Main appropriation")
        self.assertEqual(transformed_rows[0]["AmountKind"], "Total")
        self.assertEqual(transformed_rows[0]["Value"], 70000)

    def test_transforms_aene_tabular_rows_into_expected_schema(self):
        workbook_bytes = build_aene_tabular_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data (Tabular)")
        records = list(iter_sheet_records(rows, find_header_row(rows)))

        transformed_rows = transform_records(
            records,
            Mock(slug="2025-26"),
            DatasetPreparationJob.DATASET_TYPE_AENE,
        )

        self.assertEqual(list(transformed_rows[0].keys()), AENE_HEADERS)
        self.assertEqual(transformed_rows[0]["FinancialYear"], "2025")
        self.assertEqual(transformed_rows[0]["BudgetPhase"], "Adjusted appropriation")
        self.assertEqual(transformed_rows[0]["AmountKind"], "Voted")
        self.assertEqual(transformed_rows[0]["Value"], 100000)

    def test_transforms_aene_flat_rows_into_expected_schema(self):
        workbook_bytes = build_aene_flat_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data (Flat)")
        records = list(iter_sheet_records(rows, find_header_row(rows)))

        transformed_rows = transform_records(
            records,
            Mock(slug="2025-26"),
            DatasetPreparationJob.DATASET_TYPE_AENE,
        )

        self.assertEqual(len(transformed_rows), 2)
        adjusted_row = [
            row
            for row in transformed_rows
            if row["BudgetPhase"] == "Adjusted appropriation"
        ][0]
        self.assertEqual(list(adjusted_row.keys()), AENE_HEADERS)
        self.assertEqual(adjusted_row["FinancialYear"], "2025")
        self.assertEqual(adjusted_row["AmountKind"], "Voted")
        self.assertEqual(adjusted_row["Value"], 100000)

    def test_transforms_national_budget_vs_actual_rows_into_expected_schema(self):
        workbook_bytes = build_national_source_workbook()
        rows = load_rows_from_excel(workbook_bytes, "Data")
        header_row_index = find_header_row(rows)
        records = list(iter_sheet_records(rows, header_row_index))

        financial_year = Mock(slug="2026-27")
        transformed_rows = transform_records(
            records,
            financial_year,
            DatasetImportJob.DATASET_TYPE_BVA_NATIONAL,
        )

        adjusted_row = [
            row for row in transformed_rows
            if row["BudgetPhase"] == "Adjusted appropriation"
        ][0]
        main_rows = [
            row for row in transformed_rows
            if row["BudgetPhase"] == "Main appropriation"
        ]

        self.assertEqual(list(adjusted_row.keys()), BUDGET_ACTUAL_HEADERS)
        self.assertEqual(adjusted_row["Government"], "National")
        self.assertEqual(adjusted_row["AmountKind"], "Total")
        self.assertEqual(adjusted_row["Value"], 40000)
        self.assertTrue(any(row["FinancialYear"] == "2026" and row["Value"] == 70000 for row in main_rows))


class DatasetPreparationJobTest(TestCase):
    def create_common_budgetportal_setup(self, financial_year, is_provincial):
        sphere_name = "Provincial" if is_provincial else "National"
        budgetportal_models.Sphere.objects.create(
            name=sphere_name, financial_year=financial_year
        )
        budgetportal_models.Organisation.objects.create(title="National Treasury")

    @patch("dataset_preparation.tasks.async_task")
    @patch("dataset_preparation.services.requests.get")
    def test_job_prepares_files_and_queues_import_jobs(self, mock_get, mock_async_task):
        user = User.objects.create_user(username="admin", password="password")
        financial_year = budgetportal_models.FinancialYear.objects.create(
            slug="2026-27", published=True
        )
        self.create_common_budgetportal_setup(financial_year, is_provincial=True)
        budgetportal_models.DatasetCategory.objects.create(
            title="Estimates of Provincial Revenue and Expenditure",
            description="EPRE",
        )

        mock_response = Mock()
        mock_response.content = build_source_workbook()
        mock_response.raise_for_status = Mock()
        mock_get.return_value = mock_response
        mock_async_task.side_effect = ["epre-task", "bva-task"]

        job = DatasetPreparationJob.objects.create(
            user=user,
            source_url="https://example.com/epre.xlsx",
            dataset_type=DatasetPreparationJob.DATASET_TYPE_EPRE,
            financial_year=financial_year,
            sheet_name="Data",
        )

        budgetportal_models.DatasetCategory.objects.create(
            title="Budgeted and Actual Provincial Expenditure",
            description="Budget vs actual provincial",
        )

        result = run_dataset_preparation_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, DatasetPreparationJob.STATUS_PREPARED)
        self.assertTrue(job.prepared_file.name.endswith(".csv"))
        self.assertTrue(job.budget_vs_actual_file.name.endswith(".csv"))
        self.assertEqual(
            job.excel_conversion_status,
            DatasetPreparationJob.CONVERSION_STATUS_COMPLETED,
        )
        self.assertTrue(job.prepared_excel_file.name.endswith(".xlsx"))
        self.assertTrue(job.budget_vs_actual_excel_file.name.endswith(".xlsx"))
        self.assertEqual(result["epre_prepared_rows"], 1)
        self.assertEqual(result["budget_vs_actual_prepared_rows"], 1)
        self.assertEqual(job.import_jobs.count(), 2)
        self.assertEqual(
            set(job.import_jobs.values_list("status", flat=True)),
            {DatasetImportJob.STATUS_QUEUED},
        )
        self.assertEqual(
            set(job.import_jobs.values_list("task_id", flat=True)),
            {"epre-task", "bva-task"},
        )

    @patch("dataset_preparation.tasks.async_task")
    @patch("dataset_preparation.services.requests.get")
    def test_import_job_runs_separately_after_preparation(
        self, mock_get, mock_async_task
    ):
        user = User.objects.create_user(username="admin", password="password")
        financial_year = budgetportal_models.FinancialYear.objects.create(
            slug="2026-27", published=True
        )
        self.create_common_budgetportal_setup(financial_year, is_provincial=True)
        budgetportal_models.DatasetCategory.objects.create(
            title="Estimates of Provincial Revenue and Expenditure",
            description="EPRE",
        )
        budgetportal_models.DatasetCategory.objects.create(
            title="Budgeted and Actual Provincial Expenditure",
            description="Budget vs actual provincial",
        )

        mock_response = Mock()
        mock_response.content = build_source_workbook()
        mock_response.raise_for_status = Mock()
        mock_get.return_value = mock_response
        mock_async_task.side_effect = ["epre-task", "bva-task"]

        job = DatasetPreparationJob.objects.create(
            user=user,
            source_url="https://example.com/epre.xlsx",
            dataset_type=DatasetPreparationJob.DATASET_TYPE_EPRE,
            financial_year=financial_year,
            sheet_name="Data",
        )
        run_dataset_preparation_job(job.id)

        import_job = DatasetImportJob.objects.get(
            preparation_job=job,
            dataset_type=DatasetImportJob.DATASET_TYPE_EPRE,
        )
        result = run_dataset_import_job(import_job.id)

        import_job.refresh_from_db()
        self.assertEqual(import_job.status, DatasetImportJob.STATUS_COMPLETED)
        upload = budgetportal_models.DatasetUpload.objects.get(
            id=result["dataset_upload_id"]
        )
        self.assertEqual(upload.user, user)
        dataset = budgetportal_models.Dataset.objects.get(id=result["dataset_id"])
        self.assertEqual(
            dataset.title,
            "Estimates of Provincial Revenue and Expenditure 2026-27",
        )

    @patch("dataset_preparation.tasks.async_task")
    @patch("dataset_preparation.services.requests.get")
    def test_national_job_prepares_files_and_queues_import_jobs(
        self, mock_get, mock_async_task
    ):
        user = User.objects.create_user(username="admin", password="password")
        financial_year = budgetportal_models.FinancialYear.objects.create(
            slug="2026-27", published=True
        )
        self.create_common_budgetportal_setup(financial_year, is_provincial=False)
        budgetportal_models.DatasetCategory.objects.create(
            title="Estimates of National Expenditure",
            description="ENE",
        )
        budgetportal_models.DatasetCategory.objects.create(
            title="Budgeted vs Actual National Expenditure",
            description="Budget vs actual national",
        )

        mock_response = Mock()
        mock_response.content = build_national_source_workbook()
        mock_response.raise_for_status = Mock()
        consolidated_mock_response = Mock()
        consolidated_mock_response.content = build_consolidated_source_workbook()
        consolidated_mock_response.raise_for_status = Mock()
        mock_get.side_effect = [mock_response, consolidated_mock_response]
        mock_async_task.side_effect = [
            "ene-task",
            "bva-national-task",
            "consolidation-task",
        ]

        job = DatasetPreparationJob.objects.create(
            user=user,
            source_url="https://example.com/ene.xlsx",
            consolidation_source_url="https://example.com/consolidated.xlsx",
            dataset_type=DatasetPreparationJob.DATASET_TYPE_ENE,
            financial_year=financial_year,
            sheet_name="Data",
        )

        result = run_dataset_preparation_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, DatasetPreparationJob.STATUS_PREPARED)
        self.assertIn("prepared-ene-2026-27.csv", job.prepared_file.name)
        self.assertIn(
            "prepared-budget-vs-actual-national-2026-27.csv",
            job.budget_vs_actual_file.name,
        )
        self.assertIn(
            "prepared-consolidated-expenditure-2026-27.csv",
            job.consolidated_file.name,
        )
        self.assertTrue(job.prepared_excel_file.name.endswith(".xlsx"))
        self.assertTrue(job.budget_vs_actual_excel_file.name.endswith(".xlsx"))
        self.assertTrue(job.consolidated_excel_file.name.endswith(".xlsx"))
        self.assertEqual(result["epre_prepared_rows"], 1)
        self.assertEqual(result["budget_vs_actual_prepared_rows"], 1)
        self.assertEqual(result["consolidated_prepared_rows"], 1)
        self.assertEqual(
            set(job.import_jobs.values_list("dataset_type", flat=True)),
            {
                DatasetImportJob.DATASET_TYPE_ENE,
                DatasetImportJob.DATASET_TYPE_BVA_NATIONAL,
                DatasetImportJob.DATASET_TYPE_CONSOLIDATION,
            },
        )
        self.assertEqual(
            set(job.import_jobs.values_list("task_id", flat=True)),
            {"ene-task", "bva-national-task", "consolidation-task"},
        )

    @patch("dataset_preparation.tasks.async_task")
    @patch("dataset_preparation.services.requests.get")
    def test_national_import_job_runs_separately_after_preparation(
        self, mock_get, mock_async_task
    ):
        user = User.objects.create_user(username="admin", password="password")
        financial_year = budgetportal_models.FinancialYear.objects.create(
            slug="2026-27", published=True
        )
        self.create_common_budgetportal_setup(financial_year, is_provincial=False)
        budgetportal_models.DatasetCategory.objects.create(
            title="Estimates of National Expenditure",
            description="ENE",
        )
        budgetportal_models.DatasetCategory.objects.create(
            title="Budgeted vs Actual National Expenditure",
            description="Budget vs actual national",
        )

        mock_response = Mock()
        mock_response.content = build_national_source_workbook()
        mock_response.raise_for_status = Mock()
        consolidated_mock_response = Mock()
        consolidated_mock_response.content = build_consolidated_source_workbook()
        consolidated_mock_response.raise_for_status = Mock()
        mock_get.side_effect = [mock_response, consolidated_mock_response]
        mock_async_task.side_effect = [
            "ene-task",
            "bva-national-task",
            "consolidation-task",
        ]

        job = DatasetPreparationJob.objects.create(
            user=user,
            source_url="https://example.com/ene.xlsx",
            consolidation_source_url="https://example.com/consolidated.xlsx",
            dataset_type=DatasetPreparationJob.DATASET_TYPE_ENE,
            financial_year=financial_year,
            sheet_name="Data",
        )
        run_dataset_preparation_job(job.id)

        import_job = DatasetImportJob.objects.get(
            preparation_job=job,
            dataset_type=DatasetImportJob.DATASET_TYPE_ENE,
        )
        result = run_dataset_import_job(import_job.id)

        import_job.refresh_from_db()
        self.assertEqual(import_job.status, DatasetImportJob.STATUS_COMPLETED)
        upload = budgetportal_models.DatasetUpload.objects.get(
            id=result["dataset_upload_id"]
        )
        self.assertEqual(upload.user, user)
        dataset = budgetportal_models.Dataset.objects.get(id=result["dataset_id"])
        self.assertEqual(
            dataset.title,
            "Estimates of National Expenditure 2026-27",
        )
        self.assertEqual(
            set(dataset.resources.values_list("format", flat=True)), {"CSV", "XLSX"}
        )

        budget_vs_actual_import_job = DatasetImportJob.objects.get(
            preparation_job=job,
            dataset_type=DatasetImportJob.DATASET_TYPE_BVA_NATIONAL,
        )
        budget_vs_actual_result = run_dataset_import_job(budget_vs_actual_import_job.id)
        budget_vs_actual_dataset = budgetportal_models.Dataset.objects.get(
            id=budget_vs_actual_result["dataset_id"]
        )
        self.assertEqual(
            set(budget_vs_actual_dataset.resources.values_list("format", flat=True)),
            {"CSV", "XLSX"},
        )

        consolidation_import_job = DatasetImportJob.objects.get(
            preparation_job=job,
            dataset_type=DatasetImportJob.DATASET_TYPE_CONSOLIDATION,
        )
        consolidation_result = run_dataset_import_job(consolidation_import_job.id)
        consolidated_dataset = budgetportal_models.Dataset.objects.get(
            id=consolidation_result["dataset_id"]
        )
        self.assertEqual(
            consolidated_dataset.title,
            "Consolidated Expenditure 2026-27",
        )
        self.assertEqual(
            set(consolidated_dataset.resources.values_list("format", flat=True)),
            {"CSV", "XLSX"},
        )

    @patch("dataset_preparation.tasks.async_task")
    @patch("dataset_preparation.services.requests.get")
    def test_aene_job_prepares_one_file_and_queues_one_import_job(
        self, mock_get, mock_async_task
    ):
        user = User.objects.create_user(username="admin", password="password")
        financial_year = budgetportal_models.FinancialYear.objects.create(
            slug="2026-27", published=True
        )
        self.create_common_budgetportal_setup(financial_year, is_provincial=False)
        budgetportal_models.DatasetCategory.objects.create(
            title="Adjusted Estimates of National Expenditure",
            description="AENE",
        )

        mock_response = Mock()
        mock_response.content = build_national_source_workbook()
        mock_response.raise_for_status = Mock()
        mock_get.return_value = mock_response
        mock_async_task.return_value = "aene-task"

        job = DatasetPreparationJob.objects.create(
            user=user,
            source_url="https://example.com/aene.xlsx",
            dataset_type=DatasetPreparationJob.DATASET_TYPE_AENE,
            financial_year=financial_year,
            sheet_name="Data",
        )

        result = run_dataset_preparation_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, DatasetPreparationJob.STATUS_PREPARED)
        self.assertIn("prepared-aene-2026-27.csv", job.prepared_file.name)
        self.assertFalse(job.budget_vs_actual_file)
        self.assertEqual(result["epre_prepared_rows"], 1)
        self.assertIsNone(result["budget_vs_actual_prepared_rows"])
        self.assertEqual(
            list(job.import_jobs.values_list("dataset_type", flat=True)),
            [DatasetImportJob.DATASET_TYPE_AENE],
        )
        self.assertEqual(
            list(job.import_jobs.values_list("task_id", flat=True)), ["aene-task"]
        )

        import_job = job.import_jobs.get(
            dataset_type=DatasetImportJob.DATASET_TYPE_AENE
        )
        import_result = run_dataset_import_job(import_job.id)
        dataset = budgetportal_models.Dataset.objects.get(id=import_result["dataset_id"])
        self.assertEqual(
            dataset.title,
            "Adjusted Estimates of National Expenditure 2026-27",
        )
