import os
import re
import time
from csv import DictWriter
import csv
from io import BytesIO, StringIO
from urllib.parse import urlparse

import requests
from django.core.files.base import ContentFile
from openpyxl import Workbook, load_workbook

from budgetportal import models as budgetportal_models
from budgetportal.dataset_uploading import check_input_column_order
from budgetportal.dataset_uploading.dataset_preprocessor import (
    AENE_HEADERS,
    BUDGET_ACTUAL_HEADERS,
    CONSOLIDATED_HEADERS,
    ENE_HEADERS,
    EPRE_HEADERS,
)

from .models import DatasetPreparationJob


PROVINCIAL_YEAR_PHASE_REGEX = re.compile(r"^\s*(20\d{2})/\d{2}\s+(.+?)\s*$")
NATIONAL_YEAR_PHASE_REGEX = re.compile(r"^\s*(20\d{2})(?:/\d{2})?\s+(.+?)\s*$")
AENE_FLAT_YEAR_PHASE_REGEX = re.compile(r"^\s*(20\d{2})/\d{2}\s*-\s*(.+?)\s*$")

HEADER_ALIASES = {
    "province": "Government",
    "government": "Government",
    "vote no.": "VoteNumber",
    "vote no": "VoteNumber",
    "voteno": "VoteNumber",
    "vote number": "VoteNumber",
    "vote": "VoteNumber",
    "department": "Department",
    "programme no.": "ProgNumber",
    "programme no": "ProgNumber",
    "program no.": "ProgNumber",
    "prognumber": "ProgNumber",
    "programme": "Programme",
    "subprogramme no.": "SubprogNumber",
    "subprogramme no": "SubprogNumber",
    "sub program no.": "SubprogNumber",
    "subprognumber": "SubprogNumber",
    "subprogramme": "Subprogramme",
    "econ1": "EconomicClassification1",
    "econ2": "EconomicClassification2",
    "econ3": "EconomicClassification3",
    "econ4": "EconomicClassification4",
    "econ5": "EconomicClassification5",
    "economicclassification1": "EconomicClassification1",
    "economicclassification2": "EconomicClassification2",
    "economicclassification3": "EconomicClassification3",
    "economicclassification4": "EconomicClassification4",
    "economicclassification5": "EconomicClassification5",
    "economic level 1": "EconomicClassification1",
    "economic level 2": "EconomicClassification2",
    "economic level 3": "EconomicClassification3",
    "economic level 4": "EconomicClassification4",
    "economic level 5": "EconomicClassification5",
    "function group": "FunctionGroup1",
    "function group 1": "FunctionGroup1",
    "functiongroup1": "FunctionGroup1",
    "function group 2": "FunctionGroup2",
    "functiongroup2": "FunctionGroup2",
    "voted directcharges": "AmountKind",
    "voted direct charges": "AmountKind",
    "finyear": "SourceFinancialYear",
    "financial year": "SourceFinancialYear",
    "fy description": "SourceBudgetPhase",
}

TITLE_CASE_SMALL_WORDS = {
    "And": "and",
    "Of": "of",
    "The": "the",
}

PHASE_ALIASES = {
    "main appropriation": "Main appropriation",
    "voted main appropriation": "Main appropriation",
    "budget": "Main appropriation",
    "mtef1": "Main appropriation",
    "mtef 1": "Main appropriation",
    "mtef2": "Main appropriation",
    "mtef 2": "Main appropriation",
    "adjusted appropriation": "Adjusted appropriation",
    "baseline": "Baseline",
    "revised baseline": "Revised baseline",
    "revised estimate": "Revised estimate",
    "audited outcome": "Audited outcome",
    "audit outcome": "Audited outcome",
    "audited outcomes": "Audited outcome",
    "preliminary outcome": "Preliminary outcome",
    "final appropriation": "Final appropriation",
}


class DatasetPreparationError(Exception):
    pass


DOWNLOAD_MAX_ATTEMPTS = 4
DOWNLOAD_RETRY_DELAY_SECONDS = 2

PREPARATION_DATASET_CONFIG = {
    DatasetPreparationJob.DATASET_TYPE_EPRE: {
        "primary_dataset_type": DatasetPreparationJob.DATASET_TYPE_EPRE,
        "budget_vs_actual_dataset_type": "Budget-vs-Actual-Provincial",
        "primary_filename_slug": "epre",
        "budget_vs_actual_filename_slug": "budget-vs-actual-provincial",
        "primary_log_label": "EPRE",
        "budget_vs_actual_log_label": "Budget vs Actual Provincial",
    },
    DatasetPreparationJob.DATASET_TYPE_ENE: {
        "primary_dataset_type": DatasetPreparationJob.DATASET_TYPE_ENE,
        "budget_vs_actual_dataset_type": "Budget-vs-Actual-National",
        "primary_filename_slug": "ene",
        "budget_vs_actual_filename_slug": "budget-vs-actual-national",
        "primary_log_label": "ENE",
        "budget_vs_actual_log_label": "Budget vs Actual National",
        "consolidation_dataset_type": "Consolidation",
        "consolidation_filename_slug": "consolidated-expenditure",
        "consolidation_log_label": "Consolidated Expenditure",
    },
    DatasetPreparationJob.DATASET_TYPE_AENE: {
        "primary_dataset_type": DatasetPreparationJob.DATASET_TYPE_AENE,
        "primary_filename_slug": "aene",
        "primary_log_label": "AENE",
        "fallback_sheet_name": "Data (Tabular)",
    },
}


def emit_service_progress(message):
    print("[dataset_preparation.services] {}".format(message), flush=True)


def get_preparation_dataset_config(dataset_type):
    config = PREPARATION_DATASET_CONFIG.get(dataset_type)
    if not config:
        raise DatasetPreparationError(
            "Unsupported preparation type '{}'.".format(dataset_type)
        )
    return config


def download_source_file(source_url):
    emit_service_progress("Downloading source file from {}".format(source_url))
    last_error = None

    for attempt in range(1, DOWNLOAD_MAX_ATTEMPTS + 1):
        try:
            response = requests.get(source_url, timeout=120)
            response.raise_for_status()
            emit_service_progress(
                "Downloaded {} bytes from {}".format(len(response.content), source_url)
            )
            return response.content
        except requests.exceptions.HTTPError:
            # HTTP errors are not transient DNS issues, so fail fast.
            raise
        except requests.exceptions.RequestException as exc:
            last_error = exc
            if attempt == DOWNLOAD_MAX_ATTEMPTS:
                break
            emit_service_progress(
                "Download attempt {} of {} failed: {}. Retrying in {} seconds.".format(
                    attempt,
                    DOWNLOAD_MAX_ATTEMPTS,
                    exc,
                    DOWNLOAD_RETRY_DELAY_SECONDS,
                )
            )
            time.sleep(DOWNLOAD_RETRY_DELAY_SECONDS)

    raise last_error


def filename_from_url(source_url):
    path = urlparse(source_url).path
    filename = os.path.basename(path)
    return filename or "dataset-source.xlsx"


def save_content_to_field(instance, field_name, filename, content_bytes):
    content_file = ContentFile(content_bytes)
    getattr(instance, field_name).save(filename, content_file, save=False)


def load_rows_from_excel(content_bytes, sheet_name, fallback_sheet_name=None):
    emit_service_progress("Loading workbook sheet '{}'".format(sheet_name))
    workbook = load_workbook(BytesIO(content_bytes), data_only=True, read_only=True)
    target_sheet_name = sheet_name or workbook.sheetnames[0]
    if target_sheet_name not in workbook.sheetnames and fallback_sheet_name in workbook.sheetnames:
        emit_service_progress(
            "Sheet '{}' was not found; using '{}' instead.".format(
                target_sheet_name, fallback_sheet_name
            )
        )
        target_sheet_name = fallback_sheet_name
    if target_sheet_name not in workbook.sheetnames:
        raise DatasetPreparationError(
            "Could not find sheet '{}' in source workbook.".format(target_sheet_name)
        )
    worksheet = workbook[target_sheet_name]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    emit_service_progress(
        "Loaded {} rows from sheet '{}'".format(len(rows), target_sheet_name)
    )
    return rows


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", " ", normalize_text(value).lower()).strip()


def normalize_phase_name(value, dataset_type):
    normalized = normalize_key(value)
    if dataset_type in (
        DatasetPreparationJob.DATASET_TYPE_EPRE,
        "Budget-vs-Actual-Provincial",
    ):
        if normalized in {"budget", "voted main appropriation", "mtef1", "mtef 1", "mtef2", "mtef 2"}:
            return None
    return PHASE_ALIASES.get(normalized)


def get_year_phase_regex(dataset_type):
    if dataset_type in (
        DatasetPreparationJob.DATASET_TYPE_EPRE,
        "Budget-vs-Actual-Provincial",
    ):
        return PROVINCIAL_YEAR_PHASE_REGEX
    return NATIONAL_YEAR_PHASE_REGEX


def find_header_row(rows):
    for index, row in enumerate(rows):
        row_values = [normalize_key(value) for value in row]
        if "programme" in row_values:
            emit_service_progress("Found header row at index {}".format(index))
            return index
    raise DatasetPreparationError("Could not find Programme header row in source workbook.")


def build_canonical_headers(header_row):
    headers = []
    for value in header_row:
        text = normalize_text(value)
        canonical = HEADER_ALIASES.get(normalize_key(text), text)
        headers.append(canonical)
    return headers


def iter_sheet_records(rows, header_row_index):
    headers = build_canonical_headers(rows[header_row_index])
    for row in rows[header_row_index + 1 :]:
        if not any(value not in (None, "") for value in row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        yield dict(zip(headers, padded))


def extract_value_columns(headers, dataset_type):
    year_phase_regex = get_year_phase_regex(dataset_type)
    value_columns = []
    for header in headers:
        if year_phase_regex.match(normalize_text(header)):
            value_columns.append(header)
    if not value_columns:
        raise DatasetPreparationError(
            "No year/phase columns matched the expected pattern YYYY/YY."
        )
    return value_columns


def coerce_vote_number(value):
    text = normalize_text(value)
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    return int(match.group(1))


def coerce_numeric_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value))) * 1000

    text = normalize_text(value).replace(",", "").replace(" ", "")
    if not text:
        return None
    try:
        return int(round(float(text))) * 1000
    except ValueError:
        return None


def format_title_case(value):
    text = normalize_text(value)
    if not text:
        return ""
    result = text.title()
    for original, replacement in TITLE_CASE_SMALL_WORDS.items():
        result = re.sub(r"\b{}\b".format(original), replacement, result)
    return result.replace("Kwazulu-Natal", "KwaZulu-Natal")


def build_row(record, financial_year, phase_name, value, dataset_type):
    government_name = format_title_case(record.get("Government"))
    if dataset_type == "Budget-vs-Actual-National" and not government_name:
        government_name = "National"

    base_row = {
        "Government": government_name,
        "VoteNumber": coerce_vote_number(record.get("VoteNumber")),
        "Department": format_title_case(record.get("Department")),
        "ProgNumber": record.get("ProgNumber"),
        "Programme": record.get("Programme"),
        "SubprogNumber": record.get("SubprogNumber"),
        "Subprogramme": record.get("Subprogramme"),
        "EconomicClassification1": record.get("EconomicClassification1"),
        "EconomicClassification2": record.get("EconomicClassification2"),
        "EconomicClassification3": record.get("EconomicClassification3"),
        "EconomicClassification4": record.get("EconomicClassification4"),
        "EconomicClassification5": record.get("EconomicClassification5"),
        "FunctionGroup1": record.get("FunctionGroup1"),
        "BudgetYear": financial_year,
        "FinancialYear": financial_year,
        "BudgetPhase": phase_name,
        "Value": value,
    }

    if dataset_type == DatasetPreparationJob.DATASET_TYPE_EPRE:
        base_row["FunctionGroup2"] = record.get("FunctionGroup2")
        ordered_headers = EPRE_HEADERS
    elif dataset_type == DatasetPreparationJob.DATASET_TYPE_ENE:
        ordered_headers = ENE_HEADERS
    elif dataset_type == DatasetPreparationJob.DATASET_TYPE_AENE:
        base_row["AmountKind"] = "Total"
        ordered_headers = AENE_HEADERS
    else:
        base_row["AmountKind"] = "Total"
        ordered_headers = BUDGET_ACTUAL_HEADERS

    return {header: base_row.get(header) for header in ordered_headers}


def transform_aene_tabular_records(records, target_financial_year):
    target_year = target_financial_year.slug.split("-")[0]
    output_rows = []
    for record in records:
        source_financial_year = normalize_text(record.get("SourceFinancialYear"))
        source_year_match = re.match(r"(20\d{2})", source_financial_year)
        if not source_year_match or source_year_match.group(1) != target_year:
            continue

        budget_phase = normalize_text(record.get("SourceBudgetPhase"))
        value = coerce_numeric_value(record.get("Value"))
        if not budget_phase or value is None:
            continue

        row = build_row(
            record,
            source_year_match.group(1),
            budget_phase,
            value,
            DatasetPreparationJob.DATASET_TYPE_AENE,
        )
        row["AmountKind"] = normalize_text(record.get("AmountKind")) or "Total"
        output_rows.append(row)

    if not output_rows:
        raise DatasetPreparationError(
            "AENE tabular source did not contain rows for financial year {}.".format(
                target_financial_year.slug
            )
        )

    emit_service_progress("Transformed {} output rows for AENE".format(len(output_rows)))
    return output_rows


def transform_aene_flat_records(records, target_financial_year):
    target_year = target_financial_year.slug.split("-")[0]
    value_columns = [
        header
        for header in records[0]
        if AENE_FLAT_YEAR_PHASE_REGEX.match(normalize_text(header))
    ]
    output_rows = []
    for record in records:
        for column_name in value_columns:
            match = AENE_FLAT_YEAR_PHASE_REGEX.match(normalize_text(column_name))
            if not match or match.group(1) != target_year:
                continue

            value = coerce_numeric_value(record.get(column_name))
            if value is None:
                continue

            row = build_row(
                record,
                match.group(1),
                match.group(2).strip(),
                value,
                DatasetPreparationJob.DATASET_TYPE_AENE,
            )
            row["AmountKind"] = normalize_text(record.get("AmountKind")) or "Total"
            output_rows.append(row)

    if not output_rows:
        raise DatasetPreparationError(
            "AENE flat source did not contain rows for financial year {}.".format(
                target_financial_year.slug
            )
        )

    emit_service_progress("Transformed {} output rows for AENE".format(len(output_rows)))
    return output_rows


def build_consolidated_rows(ene_rows):
    """Aggregate ENE rows for legacy callers.

    New ENE preparation jobs use ``transform_consolidated_records`` with the
    Treasury consolidated-account workbook. ENE excludes provincial and social
    security expenditure, so it must not be used as the source of a new
    consolidated expenditure dataset.
    """
    totals = {}
    for row in ene_rows:
        key = (
            normalize_text(row.get("FunctionGroup1")),
            normalize_text(row.get("EconomicClassification2")),
            normalize_text(row.get("EconomicClassification3")),
            normalize_text(row.get("FinancialYear")),
        )
        totals[key] = totals.get(key, 0) + (row.get("Value") or 0)

    output_rows = [
        {
            "FunctionGroup": function_group,
            "EconomicClassification2": economic_classification2,
            "EconomicClassification3": economic_classification3,
            "FinancialYear": financial_year,
            "Value": value,
        }
        for (
            function_group,
            economic_classification2,
            economic_classification3,
            financial_year,
        ), value in sorted(totals.items())
    ]
    if not output_rows:
        raise DatasetPreparationError("ENE preparation did not produce rows for consolidation.")
    emit_service_progress(
        "Aggregated {} ENE rows into {} Consolidated Expenditure rows".format(
            len(ene_rows), len(output_rows)
        )
    )
    return output_rows


def transform_consolidated_records(records, target_financial_year):
    """Transform Treasury's Consolidated account Pivot ``Data`` sheet.

    The source workbook is already at the public consolidated granularity and
    contains its own ``FinYear`` and value columns. Values are published in
    thousands of rand and are converted to rand to match ``ConsolidationData``.
    """
    target_year = target_financial_year.slug.split("-")[0]
    output_rows = []

    for record in records:
        source_financial_year = normalize_text(record.get("SourceFinancialYear"))
        source_year_match = re.match(r"(20\d{2})", source_financial_year)
        if not source_year_match or source_year_match.group(1) != target_year:
            continue

        value = coerce_numeric_value(record.get("Value"))
        if value is None:
            continue

        output_rows.append(
            {
                "FunctionGroup": normalize_text(record.get("FunctionGroup1")),
                "EconomicClassification2": normalize_text(
                    record.get("EconomicClassification2")
                ),
                "EconomicClassification3": normalize_text(
                    record.get("EconomicClassification3")
                ),
                "FinancialYear": target_year,
                "Value": value,
            }
        )

    if not output_rows:
        raise DatasetPreparationError(
            "Consolidated source did not contain rows for financial year {}.".format(
                target_financial_year.slug
            )
        )

    emit_service_progress(
        "Transformed {} Consolidated Expenditure rows".format(len(output_rows))
    )
    return output_rows


def transform_records(records, target_financial_year, dataset_type):
    records = list(records)
    if not records:
        raise DatasetPreparationError("The source workbook does not contain any data rows.")
    emit_service_progress(
        "Transforming {} records for {} {}".format(
            len(records), dataset_type, target_financial_year.slug
        )
    )

    if (
        dataset_type == DatasetPreparationJob.DATASET_TYPE_AENE
        and "SourceFinancialYear" in records[0]
        and "SourceBudgetPhase" in records[0]
    ):
        return transform_aene_tabular_records(records, target_financial_year)
    if (
        dataset_type == DatasetPreparationJob.DATASET_TYPE_AENE
        and any(
            AENE_FLAT_YEAR_PHASE_REGEX.match(normalize_text(header))
            for header in records[0]
        )
    ):
        return transform_aene_flat_records(records, target_financial_year)

    year_phase_regex = get_year_phase_regex(dataset_type)
    value_columns = extract_value_columns(records[0].keys(), dataset_type)
    target_year = target_financial_year.slug.split("-")[0]

    if dataset_type in (
        DatasetPreparationJob.DATASET_TYPE_EPRE,
        DatasetPreparationJob.DATASET_TYPE_ENE,
    ):
        wanted_phases = {
            "Baseline",
            "Revised baseline",
            "Main appropriation",
        }
    elif dataset_type == DatasetPreparationJob.DATASET_TYPE_AENE:
        wanted_phases = {
            "Adjusted appropriation",
            "Audited outcome",
            "Preliminary outcome",
            "Baseline",
            "Main appropriation",
            "Revised baseline",
            "Revised estimate",
            "Final appropriation",
        }
    elif dataset_type in (
        "Budget-vs-Actual-Provincial",
        "Budget-vs-Actual-National",
    ):
        wanted_phases = {
            "Adjusted appropriation",
            "Audited outcome",
            "Preliminary outcome",
            "Baseline",
            "Main appropriation",
            "Revised baseline",
            "Revised estimate",
            "Final appropriation",
        }
    else:
        raise DatasetPreparationError(
            "Unsupported preparation type '{}'.".format(dataset_type)
        )
    output_rows = []
    discovered_columns = []
    for record in records:
        for column_name in value_columns:
            match = year_phase_regex.match(normalize_text(column_name))
            if not match:
                continue
            source_financial_year = match.group(1)
            raw_phase_name = match.group(2).strip()
            phase_name = normalize_phase_name(raw_phase_name, dataset_type)
            discovered_columns.append("{} {}".format(source_financial_year, raw_phase_name))
            if phase_name is None:
                continue
            if phase_name not in wanted_phases:
                continue

            value = coerce_numeric_value(record.get(column_name))
            if value is None:
                continue

            if dataset_type in (
                DatasetPreparationJob.DATASET_TYPE_EPRE,
                DatasetPreparationJob.DATASET_TYPE_ENE,
                DatasetPreparationJob.DATASET_TYPE_AENE,
            ):
                if source_financial_year != target_year:
                    continue
                if dataset_type == DatasetPreparationJob.DATASET_TYPE_AENE:
                    budget_phase = phase_name
                else:
                    budget_phase = "Main appropriation"
            else:
                budget_phase = phase_name

            output_rows.append(
                build_row(record, source_financial_year, budget_phase, value, dataset_type)
            )

    if not output_rows:
        raise DatasetPreparationError(
            "Preparation completed but no rows matched the expected phases for "
            "{} in financial year {}. Found year/phase columns: {}.".format(
                dataset_type,
                target_financial_year.slug,
                ", ".join(sorted(set(discovered_columns))) or "none",
            )
        )

    emit_service_progress(
        "Transformed {} output rows for {}".format(len(output_rows), dataset_type)
    )
    return output_rows


def write_prepared_csv(rows, headers):
    emit_service_progress("Writing {} rows to CSV".format(len(rows)))
    buffer = StringIO()
    writer = DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header) for header in headers})
    return buffer.getvalue().encode("utf-8-sig")


def read_csv_bytes(content_bytes):
    decoded = content_bytes.decode("utf-8-sig")
    rows = decoded.splitlines()
    if not rows:
        raise DatasetPreparationError("Prepared CSV file is empty.")
    import csv

    reader = csv.reader(rows)
    rows = list(reader)
    if not rows:
        raise DatasetPreparationError("Prepared CSV file is empty.")
    return rows[0], rows[1:]


def write_excel_from_csv_bytes(content_bytes):
    # Stream the CSV into the write-only workbook. Budget-vs-Actual files can
    # contain more than 170,000 rows; materialising every parsed row first can
    # exhaust the worker and leave the preparation job stuck at this step.
    csv_stream = StringIO(content_bytes.decode("utf-8-sig"))
    reader = csv.reader(csv_stream)
    headers = next(reader, None)
    if not headers:
        raise DatasetPreparationError("Prepared CSV file is empty.")

    emit_service_progress("Converting CSV to Excel")
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="Prepared Dataset")
    worksheet.append(headers)
    row_count = 0
    for row in reader:
        worksheet.append(row)
        row_count += 1
    buffer = BytesIO()
    workbook.save(buffer)
    emit_service_progress(
        "Converted CSV to Excel with {} data rows".format(row_count)
    )
    return buffer.getvalue()


def validate_prepared_headers(headers, dataset_type):
    if dataset_type == DatasetPreparationJob.DATASET_TYPE_EPRE:
        expected_headers = EPRE_HEADERS
    elif dataset_type == DatasetPreparationJob.DATASET_TYPE_ENE:
        expected_headers = ENE_HEADERS
    elif dataset_type == DatasetPreparationJob.DATASET_TYPE_AENE:
        expected_headers = AENE_HEADERS
    elif dataset_type == "Consolidation":
        expected_headers = CONSOLIDATED_HEADERS
    else:
        expected_headers = BUDGET_ACTUAL_HEADERS
    check_input_column_order(headers, expected_headers, dataset_type)


def prepare_dataset_job(job):
    emit_service_progress(
        "Preparing dataset job {} for financial year {}".format(
            job.id, job.financial_year.slug
        )
    )
    source_content = download_source_file(job.source_url)
    raw_filename = filename_from_url(job.source_url)
    save_content_to_field(job, "raw_file", raw_filename, source_content)
    emit_service_progress("Saved raw file as {}".format(raw_filename))

    config = get_preparation_dataset_config(job.dataset_type)
    rows = load_rows_from_excel(
        source_content,
        job.sheet_name,
        fallback_sheet_name=config.get("fallback_sheet_name"),
    )
    header_row_index = find_header_row(rows)
    records = list(iter_sheet_records(rows, header_row_index))

    primary_rows = transform_records(
        records,
        job.financial_year,
        config["primary_dataset_type"],
    )
    validate_prepared_headers(
        list(primary_rows[0].keys()), config["primary_dataset_type"]
    )
    primary_bytes = write_prepared_csv(primary_rows, list(primary_rows[0].keys()))
    save_content_to_field(
        job,
        "prepared_file",
        "prepared-{}-{}.csv".format(
            config["primary_filename_slug"],
            job.financial_year.slug,
        ),
        primary_bytes,
    )
    emit_service_progress(
        "Saved {} prepared file with {} rows".format(
            config["primary_log_label"], len(primary_rows)
        )
    )

    results = {config["primary_dataset_type"]: primary_rows}
    if config.get("budget_vs_actual_dataset_type"):
        budget_vs_actual_rows = transform_records(
            records,
            job.financial_year,
            config["budget_vs_actual_dataset_type"],
        )
        validate_prepared_headers(
            list(budget_vs_actual_rows[0].keys()),
            config["budget_vs_actual_dataset_type"],
        )
        budget_vs_actual_bytes = write_prepared_csv(
            budget_vs_actual_rows,
            list(budget_vs_actual_rows[0].keys()),
        )
        save_content_to_field(
            job,
            "budget_vs_actual_file",
            "prepared-{}-{}.csv".format(
                config["budget_vs_actual_filename_slug"],
                job.financial_year.slug,
            ),
            budget_vs_actual_bytes,
        )
        emit_service_progress(
            "Saved {} prepared file with {} rows".format(
                config["budget_vs_actual_log_label"],
                len(budget_vs_actual_rows),
            )
        )
        results[config["budget_vs_actual_dataset_type"]] = budget_vs_actual_rows

    if config.get("consolidation_dataset_type"):
        if not job.consolidation_source_url:
            raise DatasetPreparationError(
                "An ENE preparation job needs a Consolidated account source URL. "
                "Provide the National Treasury Consolidated account Pivot workbook."
            )
        consolidated_source_content = download_source_file(job.consolidation_source_url)
        consolidated_rows_source = load_rows_from_excel(
            consolidated_source_content,
            "Data",
        )
        consolidated_records = list(iter_sheet_records(consolidated_rows_source, 0))
        consolidated_rows = transform_consolidated_records(
            consolidated_records,
            job.financial_year,
        )
        validate_prepared_headers(
            list(consolidated_rows[0].keys()), config["consolidation_dataset_type"]
        )
        consolidated_bytes = write_prepared_csv(consolidated_rows, CONSOLIDATED_HEADERS)
        save_content_to_field(
            job,
            "consolidated_file",
            "prepared-{}-{}.csv".format(
                config["consolidation_filename_slug"], job.financial_year.slug
            ),
            consolidated_bytes,
        )
        emit_service_progress(
            "Saved {} prepared file with {} rows".format(
                config["consolidation_log_label"], len(consolidated_rows)
            )
        )
        results[config["consolidation_dataset_type"]] = consolidated_rows

    return results


def create_dataset_upload_from_field(job, field_name, dataset_type):
    prepared_field = getattr(job, field_name)
    if not prepared_field:
        raise DatasetPreparationError(
            "Prepared file '{}' is missing for this job.".format(field_name)
        )

    upload = budgetportal_models.DatasetUpload(
        user=job.user,
        type=dataset_type,
        financialYear=job.financial_year,
    )
    prepared_field.open("rb")
    upload.file.save(
        os.path.basename(prepared_field.name),
        ContentFile(prepared_field.read()),
        save=False,
    )
    prepared_field.close()
    upload.save()
    return upload


def convert_prepared_csvs_to_excel(job):
    conversions = [("prepared_file", "prepared_excel_file")]
    if job.budget_vs_actual_file:
        conversions.append(("budget_vs_actual_file", "budget_vs_actual_excel_file"))
    if job.consolidated_file:
        conversions.append(("consolidated_file", "consolidated_excel_file"))
    for source_field_name, target_field_name in conversions:
        source_field = getattr(job, source_field_name)
        if not source_field:
            raise DatasetPreparationError(
                "Prepared CSV '{}' is missing for this job.".format(source_field_name)
            )
        emit_service_progress(
            "Converting {} for job {}".format(source_field.name, job.id)
        )
        source_field.open("rb")
        excel_bytes = write_excel_from_csv_bytes(source_field.read())
        source_field.close()
        target_name = os.path.splitext(os.path.basename(source_field.name))[0] + ".xlsx"
        save_content_to_field(job, target_field_name, target_name, excel_bytes)
        emit_service_progress("Saved Excel copy as {}".format(target_name))
